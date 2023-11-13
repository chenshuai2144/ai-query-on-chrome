import os
import openpyxl

# 指定文件夹路径
folder_path = "C:\\Users\\win10\\Documents\\业务答疑数据"  # 替换为实际的文件夹路径

# 遍历文件夹中的所有文件
for filename in os.listdir(folder_path):
    if (
        filename.endswith(".xlsx")
        or filename.endswith(".xls")
        or filename.endswith(".csv")
    ):  # 只处理Excel文件
        file_path = os.path.join(folder_path, filename)
        workbook = openpyxl.load_workbook(file_path)

        md = ""

        for sheetname in workbook.sheetnames:
            worksheet = workbook[sheetname]

            for row in worksheet.iter_rows(values_only=True):
                # 在这里处理每一行的数据
                index = 0
                for cell in row:
                    if index == 0:
                        md = md + "# " + str(cell) + "\n\n"
                    else:
                        print(cell)
                        if cell is None:
                            md = ""
                        else:
                            md = md + str(cell) + "\n\n\n"
                    index += 1

            # 关闭Excel文件
            workbook.close()

        # 将md字符串写入文件
        with open(
            "faq_docs" + "/" + os.path.basename(filename).split(".")[0] + ".md",
            "w",
            encoding="utf-8",
        ) as file:
            file.write(md)
